from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
from pathlib import Path
from datetime import date
import pytest
import openpyxl
from constants import test_sauceConstants as tsc

# 3A --> Act, Arrange, Assert
class Test_Demo:
    # Will call before each test.
    def setup_method(self): 
        self.driver = webdriver.Safari()
        self.driver.maximize_window()
        self.driver.get(tsc.url)
        self.find_user_name_and_password()
        # Take the date of today and check is there any file if there is not, create.
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)
    # Will call after each test.
    def teardown_method(self):
        self.driver.quit()
    
    def find_user_name_and_password(self):
        self.wait_for_element_visible((By.ID, "user-name"))
        self.user_name = self.driver.find_element(By.ID, "user-name")
        self.wait_for_element_visible((By.ID, "password"))
        self.password = self.driver.find_element(By.ID, "password")

    def wait_for_element_visible(self, locator, timeout = 5):
        WebDriverWait(self.driver, timeout).until(ec.visibility_of_element_located(locator))
    
    def wait_for_element_clickable(self, locator, timeout = 5):
        WebDriverWait(self.driver, timeout).until(ec.element_to_be_clickable(locator))

    def create_action_chain(self, username, password):
        valid_login = ActionChains(self.driver)
        valid_login.send_keys_to_element(self.user_name, username)
        valid_login.send_keys_to_element(self.password, password)
        valid_login.perform()

    def get_data():
        excel_file = openpyxl.load_workbook("data/invalid_login.xls")
        selected_sheet = excel_file["invalid_login"]

        data = list()
        total_rows = selected_sheet.max_row
        for var in range(2, total_rows + 1):
            username = selected_sheet.cell(var, 1).value
            password = selected_sheet.cell(var, 2).value
            tuple_data = (username, password)
            data.append(tuple_data)

        return data

    def save_screenshot(self, name):
        self.driver.save_screenshot(self.folderPath + "/" + name)
    
    def test_valid_login(self):
        self.create_action_chain("standard_user", "secret_sauce")
        self.wait_for_element_clickable((By.CLASS_NAME, "submit-button"))
        login_btn = self.driver.find_element(By.CLASS_NAME, "submit-button")
        login_btn.click()
        list_item = self.driver.find_elements(By.CLASS_NAME, "inventory_item")
        self.save_screenshot("test-valid-login.png")
        assert len(list_item) == 6
    
    def test_invalid_login(self):
        self.create_action_chain("test", "test123")
        self.wait_for_element_clickable((By.CLASS_NAME, "submit-button"))
        login_btn = self.driver.find_element(By.CLASS_NAME, "submit-button")
        login_btn.click()
        self.wait_for_element_visible((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3"))
        error_message = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        self.save_screenshot("test-invalid-login.png")
        assert error_message.text == "Epic sadface: Username and password do not match any user in this service"
    
    def test_login_empty(self):
        self.create_action_chain("","")
        self.wait_for_element_clickable((By.CLASS_NAME, "submit-button"))
        login_btn = self.driver.find_element(By.CLASS_NAME, "submit-button")
        login_btn.click()
        error_message = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3/text()")
        self.save_screenshot("test-login-empty.png")
        assert error_message.text == "Epic sadface: Username is required"

    def test_login_empty_password(self):
        self.create_action_chain("test", "")
        self.wait_for_element_clickable((By.CLASS_NAME, "submit-button"))
        login_button = self.driver.find_element(By.CLASS_NAME, "submit-button")
        login_button.click()
        self.wait_for_element_visible((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3"))
        error_message = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        self.save_screenshot("test-login-empty-password.png")
        assert error_message.text == "Epic sadface: Password is required"
    
    def test_login_locked_out(self):
        self.create_action_chain("locked_out_user", "secret_sauce")
        self.wait_for_element_clickable((By.CLASS_NAME, "submit-button"))
        login_button = self.driver.find_element(By.CLASS_NAME, "submit-button")
        login_button.click()
        self.wait_for_element_visible((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3"))
        error_message = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3")
        self.save_screenshot("test-login-locked-out.png")
        assert error_message.text == "Epic sadface: Sorry, this user has been locked out."

    def test_login_empty_2(self):
        self.create_action_chain("", "")
        self.wait_for_element_clickable((By.CLASS_NAME, "submit-button"))
        login_btn = self.driver.find_element(By.CLASS_NAME, "submit-button")
        login_btn.click()
        self.wait_for_element_visible((By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3/button"))
        error_button = self.driver.find_element(By.XPATH, "//*[@id='login_button_container']/div/form/div[3]/h3/button")
        error_button.click()
        self.save_screenshot("test-login-empty-2.png")
        assert True

    def test_add_product(self):
        self.create_action_chain("standard_user", "secret_sauce")
        self.wait_for_element_clickable((By.CLASS_NAME, "submit-button"))
        login_btn = self.driver.find_element(By.CLASS_NAME, "submit-button")
        login_btn.click()
        self.wait_for_element_clickable((By.XPATH, "//*[@id='add-to-cart-sauce-labs-bike-light']"))
        add_button = self.driver.find_element(By.XPATH, "//*[@id='add-to-cart-sauce-labs-bike-light']")
        add_button.click()
        self.wait_for_element_clickable((By.CLASS_NAME, "shopping_cart_link"))
        cart_button = self.driver.find_element(By.CLASS_NAME, "shopping_cart_link")
        cart_button.click()
        self.wait_for_element_visible((By.CLASS_NAME, "shopping_cart_badge"))
        test_cart = self.driver.find_element(By.CLASS_NAME, "shopping_cart_badge")
        self.save_screenshot("test-add-product.png")
        assert int(test_cart.text) > 0
        
    def test_remove_product(self):
        self.create_action_chain("standard_user", "secret_sauce")
        self.wait_for_element_clickable((By.CLASS_NAME, "submit-button"))
        login_btn = self.driver.find_element(By.CLASS_NAME, "submit-button")
        login_btn.click()
        self.wait_for_element_visible((By.XPATH, "//*[@id='add-to-cart-sauce-labs-bike-light']"))
        add_button = self.driver.find_element(By.XPATH, "//*[@id='add-to-cart-sauce-labs-bike-light']")
        add_button.click()
        self.wait_for_element_visible((By.CLASS_NAME, "shopping_cart_link"))
        #add_button = self.driver.find_element(By.XPATH, "//*[@id='add-to-cart-sauce-labs-fleece-jacket']")
        #add_button.click()
        #sleep(1)
        test_cart = self.driver.find_element(By.CLASS_NAME, "shopping_cart_link")
        test_cart.click()
        self.wait_for_element_clickable((By.ID, "remove-sauce-labs-bike-light"))
        remove_button = self.driver.find_element(By.ID, "remove-sauce-labs-bike-light")
        remove_button.click()
        cart_item_count = self.driver.find_elements(By.CLASS_NAME, "cart_item")
        try: 
            for element in cart_item_count:
                print(element)
        except Exception as e:
            print(e)
        self.save_screenshot("test-remove-product.png")
        assert cart_item_count == []

    def test_control_price_list(self):
        self.create_action_chain("standard_user", "secret_sauce")
        self.wait_for_element_clickable((By.CLASS_NAME, "submit-button"))
        login_btn = self.driver.find_element(By.CLASS_NAME, "submit-button")
        login_btn.click()
        my_list = list()
        limit = self.driver.find_elements(By.CLASS_NAME, "inventory_item_price")
        width = self.driver.execute_script("return arguments[0].offsetWidth", limit) # We can take the size of defined value with that function.
        
        for element in limit:
            my_list.append(float(element.text[1 : width]))
        self.save_screenshot("test-control-price-list.png")
        assert my_list == [29.99, 9.99, 15.99, 49.99, 7.99, 15.99]